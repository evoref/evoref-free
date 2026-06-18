"""XLSX Extractor

.xlsx ファイルからシート・行のテキストを抽出する。
openpyxl を使用。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, BinaryIO

from backend.extraction._binary_source_base import BinarySourceExtractorBase
from backend.extraction.base import ExtractionError


class XlsxExtractor(BinarySourceExtractorBase):
    """XLSX ファイルからテキストを抽出"""

    @property
    def extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    @property
    def error_code(self) -> str:
        return "xlsx_error"

    @property
    def requires(self) -> list[str]:
        return ["openpyxl"]

    def is_available(self) -> bool:
        try:
            from openpyxl import load_workbook  # noqa: F401
            return True
        except ImportError:
            return False

    def _extract_from_source(
        self,
        source: Path | BinaryIO,
        source_name: str,
    ) -> tuple[str, dict[str, Any]]:
        load_workbook = self._import_openpyxl()
        # openpyxl は Path / str / file-like を受け付けるが、
        # 既存挙動 (str(path)) を維持するため Path のときのみ str に変換する
        arg = str(source) if isinstance(source, Path) else source
        wb = load_workbook(arg, read_only=True, data_only=True)
        try:
            text, sheet_names = self._extract_sheets(wb)
        finally:
            wb.close()
        return text, {"sheets": sheet_names}

    @staticmethod
    def _import_openpyxl():
        try:
            from openpyxl import load_workbook
            return load_workbook
        except ImportError:
            raise ExtractionError(
                "missing_library",
                "openpyxl is required for XLSX extraction: pip install openpyxl",
            )

    @staticmethod
    def _extract_sheets(wb) -> tuple[str, list[str]]:
        """全シートからテキストを抽出"""
        texts: list[str] = []
        sheet_names: list[str] = []
        for sheet in wb.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                texts.append(f"[Sheet: {sheet.title}]\n" + "\n".join(rows))
                sheet_names.append(sheet.title)
        return "\n\n".join(texts), sheet_names
