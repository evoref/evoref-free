"""XLSX Writer

.xlsx: テーブルデータ → Excel ワークブック
openpyxl を使用。
"""

from __future__ import annotations

import io

from backend.export._writer_base import BytesWriterBase
from backend.export.base import ExportContent, ExportError


def _extract_tables(content: ExportContent) -> list[tuple[str, list[list[str]]]]:
    """ExportContent からテーブルデータを抽出（シート名, 行データ のリスト）"""
    tables: list[tuple[str, list[list[str]]]] = []

    if content.raw_data is not None:
        if isinstance(content.raw_data, list) and content.raw_data:
            first = content.raw_data[0]
            if isinstance(first, dict):
                headers = list(first.keys())
                rows = [headers]
                for item in content.raw_data:
                    rows.append([item.get(h, "") for h in headers])
                tables.append((content.title or "Sheet1", rows))
            elif isinstance(first, (list, tuple)):
                rows = [list(row) for row in content.raw_data]
                tables.append((content.title or "Sheet1", rows))
        return tables

    # blocks 中のテーブルを収集
    sheet_idx = 1
    for block in content.blocks:
        if block.type == "table" and block.rows:
            tables.append((f"Sheet{sheet_idx}", block.rows))
            sheet_idx += 1

    return tables


def _build_xlsx(content: ExportContent) -> bytes:
    """ExportContent を XLSX バイトデータに変換"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    tables = _extract_tables(content)
    if not tables:
        raise ExportError(
            "no_table_data",
            "No table data found in content for XLSX export",
        )

    wb = Workbook()
    # デフォルトシートを削除（後で追加するため）
    wb.remove(wb.active)

    for sheet_name, rows in tables:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel シート名は31文字まで

        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_value(value))
                # ヘッダー行のスタイル
                if row_idx == 1:
                    cell.font = Font(bold=True)

        # 列幅自動調整
        for col_idx in range(1, len(rows[0]) + 1 if rows else 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_data in rows:
                if col_idx <= len(row_data):
                    cell_len = len(str(row_data[col_idx - 1]))
                    max_len = max(max_len, cell_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

        # フィルター設定（ヘッダー行）
        if rows:
            ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce_value(value: object) -> object:
    """文字列を数値・日付に型推定"""
    if not isinstance(value, str):
        return value
    # 整数
    try:
        return int(value)
    except (ValueError, TypeError):
        pass
    # 浮動小数点数
    try:
        return float(value)
    except (ValueError, TypeError):
        pass
    return value


class XlsxWriter(BytesWriterBase):
    """XLSX ファイル Writer"""

    @property
    def extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    @property
    def requires(self) -> list[str]:
        return ["openpyxl"]

    def _render_bytes(self, content: ExportContent, ext: str) -> bytes:  # noqa: ARG002
        return _build_xlsx(content)

    def is_available(self) -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False
